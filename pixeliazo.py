#!/usr/bin/python3
"""Convert images to Excel spreadsheet with color index values for kids to draw.

This script takes an image and converts it to an Excel spreadsheet with cells
filled with numbers representing the color index values of the image. Image is
recolored with a maximum of 32 standard common colors. It allows scaling the
image and provides the option to recolor it with less number of colors
(between 2 and 32 inclusive).  

The purpose of the script is to create educational coloring exercises for small
kids to understand image representation in computer science.  

The created Excel file contains 2 spreadsheets and will have the image's name
with the suffix '.xlsx'. The first spreadsheet is a grid with values
corresponding to the pixel colors and a legend with the indices and color
names.  

![First spreadsheet screenshot](screenshot1.png "First spreadsheet screenshot")

The second spreadsheet has cells with backgound colors according to
the image. Both spreadsheets' cells are square formatted.  

![Second spreadsheet screenshot](screenshot2.png "Second spreadsheet screenshot")

Dependecies
-----------
Script requires:
* Pillow (Python Imaging Library fork) and
* openpyxl (Python library to read/write Excel 2010 files)  

Install dependecies:
* pip3 install Pillow
* pip3 install openpyxl

Usage
-----
* pixeliazo.py image [--width `int`] [--colors `int`] [--lang `str`]
[--resample `str`]
* pixeliazo.py (-h | --help)

Options:  
* -h, --help  
    Show help message and exit.  
* image  
    The image filename that will be used as input.  
* -w, --width `int`  
    Force scaling to the width provided. Width must be greater than 0.  
* -c, --colors `int`  
    Recolor the image with the specified number of colors (including white).
    32 standard colors are used if omitted. Valid values are between 2 and 32
    inclusive.  
* -l, --lang `str`  
    Language to be used for the legend in the output spreadsheet. Default
    is 'en' (English).  
* -r, --resample `str`  
    Resample filter to be used while resizing. It is ignored if resizing width
    is not set or if an invalid option given. Available options: NEAREST, BOX,
    BILINEAR, HAMMING, BICUBIC or LANCZOS. Use NEAREST if image is already in
    pixel art form. Default is BICUBIC.  

Examples
--------
* pixeliazo.py funny_image.jpg  
    funny_image.jpg.xlsx will be created.
* pixeliazo.py another_image.png --width 20  
    Image will be rescaled to 20 pixels width and the proportional
    height. 32 standard common colors will be used.
* pixeliazo.py example_image.jpg --width 20 -c 5  
    Image will be rescaled and recolored with 5 common colors. A legend
    with 5 indeces (4 if white is used) and color names will be added
    to the first spreadsheet.
* pixeliazo.py awesome_image.jpg -w 20 -c 5 --lang el  
    Image will be rescaled and recolored and a legend with indeces
    and color names will be added. Color names and captions will have
    Greek names.
* pixeliazo.py pixel_image.jpg -w 20 -r NEAREST  
    Image will be rescaled using NEAREST resample filter (this is ideal if
    your image is already in pixel art form).
"""

import sys #System
import argparse #Argument parsing
from PIL import Image,ImageColor #Python Image Library
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import NamedStyle,Alignment,PatternFill,Border,Side

#List of color names in English
COLORNAMES=['White','Black','Grey','Silver',
            'Red','Lime','Blue','Cyan',
            'Yellow','Magenta','Purple','Green',
            'Maroon','Olive','Navy','Teal',
            'Orange','Brown','Pink','Chocolate',
            'Dark grey','Sky blue','Dark green','Dark magenta',
            'Gold','Coral','Dark orange','Hot pink',
            'Khaki','Dark khaki','Violet','Orange red']
#Create the standard 32 colors palette from the COLORNAMES colors
PALETTE_DATA=[]
for n in COLORNAMES:
    PALETTE_DATA.extend(ImageColor.getrgb(n.replace(" ", "")))
#Resample filter names from PIL.Image
FILTERS=['NEAREST','BOX','BILINEAR','HAMMING','BICUBIC','LANCZOS']

def get_color_index(color,palette):
    """Returns the index of color in palette.

    Parameters:
        color:    List with 3 int values (RGB)
        palette:  List with colors each represented by 3 int values (RGB)

    Returns:
        Index of color in palette.
        -1 if not found.
    """

    for x in range(0,len(palette),3):
        if palette[x:x+3]==color: return x//3
    return -1 #Not found

def get_colors_legend(image):
    """Maps each image unique color to a number starting from 1, the
    corresponding index in the COLORNAMES list (also index in PALETTE_DATA
    divided by 3) and its hexadecimal representation. White color
    [255,255,255] is ignored because it is not added to legend and
    no background color needs to be applied.

    Parameters:
        image:    A palette mode PIL.Image. Image palette is supposed to have
                  colors from the PALETTE_DATA
                  
    Returns:
        Dictionary from image's palette colors (except white) to tuples
        with 3 items:
            An index starting from 1
            The index in the COLORNAMES list that matches this color
            A String with hexadecimal representation of the color
    """

    legend={} #New dictionary
    counter=1
    palette=image.getpalette()
    #getcolors() returns tuples (number of pixels, color index)
    for _,color in image.getcolors(): 
        colorvalue=palette[3*color:3*(color+1)]
        if colorvalue==[255,255,255]: continue #Ignore white, not in legend
        #Calculate hex representation of color
        hexcolor=''
        for v in colorvalue: hexcolor+=f'{v:02X}'
        #Add to dictionary
        legend[color]=(counter,
                       get_color_index(colorvalue,PALETTE_DATA),
                       hexcolor)
        counter+=1
    return legend #Mapping done

def create_workbook(filename,image,captions,colornames):
    """Creates a workbook with the image provided, saves it and returns it.

    Creates a workbook with two spreadsheets, first consists of a grid with
    color values and a legend, second of a grid with background colors matching
    image's pixels.

    Parameters:
        filename:    Filename to be used for saving workbook.
        image:       The image to be converted to Excel spreadsheet.
        colors:      Number of colors of the image
        captions:    List with 3 caption texts for the first spreadsheet
        colornames:  List with 32 color names

    Returns:
        openpyxl.Workbook object that has been created.
        None if saving the workbook failed.
    """

    wb=Workbook() #New workbook
    ws1=wb.active #First worksheet
    ws1.title=captions[1] #Set title
    ws2=wb.create_sheet(title=captions[2]) #Create 2nd worksheet and set title
    for ws in wb.worksheets: 
        ws.page_setup.fitToPage=True #Fit to one page
        for x in range(image.width + (3 if ws==ws1 else 0)):
            #We need 3 extra columns for the legend
            #The following equals 20 pixels, default row height is the same
            ws.column_dimensions[get_column_letter(x+1)].width=2.857   
    #Write the legend
    #Get map of image palette colors to our color names
    colorlegend=get_colors_legend(image)
    for value in colorlegend.values():
        #Numbers of colors
        ws1.cell(column=image.width+2,row=value[0]+4,value=value[0])
        ws1.cell(column=image.width+3,row=value[0]+4,value='=') #Equal sign
        #Center equal signs
        ws1.cell(column=image.width+3,row=value[0]+4).alignment=Alignment(
            horizontal='center')
        #Color names
        ws1.cell(column=image.width+4,row=value[0]+4,value=colornames[value[1]])
    #Legend caption
    ws1.merge_cells(start_row=1,
                    start_column=image.width+2,
                    end_row=3,
                    end_column=image.width+6) #Merge legend caption cells
    #Set value for legend caption
    ws1.cell(row=1,column=image.width+2,value=captions[0]) 
    #Align and wrap text
    ws1.cell(row=1,column=image.width+2).alignment=Alignment(
        horizontal='left', vertical='top',wrap_text=True)
    #Fill cells with values or colors according to pixels
    #Take care of white color which is not listed in colorlegend
    whitecolorindex=get_color_index([255,255,255],image.getpalette())
    #Border for cells in first worksheet
    thin_border_style=NamedStyle(name='thin_border_style')
    sb=Side(style='thin')
    thin_border_style.border = Border(left=sb,right=sb,top=sb,bottom=sb)
    wb.add_named_style(thin_border_style)
    #Get image data and process them
    pixelnumber=0
    for color in image.getdata():
        y,x=divmod(pixelnumber,image.width)
        pixelnumber+=1
        ws1.cell(column=x+1,row=y+1).style='thin_border_style'
        if color!=whitecolorindex:
            ws1.cell(column=x+1,row=y+1).value=colorlegend[color][0]
            ws2.cell(column=x+1,row=y+1).fill=PatternFill(
                fill_type='solid',start_color=colorlegend[color][2])
    try:
        wb.save(filename=filename) #Save the workbook
    except:
        print('Error writing workbook file:',filename)
        return None
    return wb

def load_language(lang):
    """Loads language dependent texts.

    Loads the texts stored at <lang>.txt file. If lang parameter equals 'en' or
    None, then predefined English values are used.

    Parameters:
        lang:    String with language code or None to use English.

    Returns:
        A tuple with two lists: a list with captions and a list with 32 color
        names in the language that was read from file or in English if
        reading failed.        
    """
    
    #Set predefined (English) text variables
    captions=['Paint the boxes with the appropriate colors to reveal the hidden '
              'image.',
              'Draw the pixels',
              'Painted picture']
    colornames=COLORNAMES #Point to English color names    
    if lang and lang.lower()!='en': #Load color names in the language choosen
        try:
            with open(lang.lower()+'.txt', 'r') as lang_file:
                temp_list=lang_file.read().splitlines()
                temp_captions=temp_list[:3]
                temp_colornames=temp_list[3:]
            if len(temp_captions)!=3 or len(temp_colornames)!=32:
                #Propably not a valid language file.
                #Raise error and fallback to English.
                raise IOError('Not a valid language file.')
            #Loading language file done correctly. Keep read values.
            captions=temp_captions
            colornames=temp_colornames
        except:
            print('Error loading language file. English will be used.')
    return (captions,colornames) #Return a tuple with 2 lists

def process_image(image,width,colors,resample):
    """Function to load and process image file.

    Loads the input file and converts it to RGB mode if needed. Then scales
    image if choosen by the user. Reduces number of colors and maps colors to
    module's standard palette (PALETTE_DATA).
    
    Parameters:
        image:    Input image filename, String object.
        width:    The desired width to scale the image, integer greater than 0
                  or None.
        colors:   Number of colors to be used for recoloring, positive integer.
        resample: Filter to use for resizing. String object (case insensitive)
                  or None. Available options: 'NEAREST', 'BOX', 'BILINEAR',
                  'HAMMING', 'BICUBIC' or 'LANCZOS'.
                  If None or invalid, then 'BICUBIC' is used.

    Returns:
        PIL.Image object with the image scaled down and recolored as defined by
        the input parameters.
        None if error occured while loading image.
    """

    try:
        img = Image.open(image) #Read the image
    except:
        print('Error reading image file:',image)
        return None
    if img.mode!='RGB':
        img=img.convert('RGB') #Convert image to RGB mode    
    if width and width>0: #Scale if needed
        res_filter=Image.BICUBIC #Default if resample is None or invalid
        if resample: #Not None
            try:
                #AttributeError here if resample is invalid
                res_filter=eval('Image.'+resample.upper())
            except:
                print('Invalid resample filter. BICUBIC will be used')
        img=img.resize(size=(width, int(width*img.height/img.width)),
                         resample=res_filter)
    elif width:
        print('Invalid width given. Scaling will not be done.')
    #Quantize to less colors
    #Max Coverage is important here, does not mess with colors
    img=img.quantize(colors=colors,method=Image.MAXCOVERAGE)
    img=img.convert('RGB') #Convert again to RGB mode (quantize made it P mode)
    #Do recoloring to standard 32 colors palette
    palimage = Image.new('P',(16,16))
    palimage.putpalette(PALETTE_DATA+[0]*(768-len(PALETTE_DATA)))   
    #Quantize image, no dithering, we want sharp image
    img=img.quantize(palette=palimage, dither=0)
    return img #Return the PIL.Image object

def number_of_colors(value):
    """Function for parsing number of colors argument.

    Parameters:
        value:    String that contains an integer
        
    Raises:
        argparse.ArgumentTypeError if value is not integer
        or out of valid range.

    Returns:
        Integer value of value parameter
    """

    error=(f'\'{value}\' is an invalid number of '
           'colors. Use integer between 2 and '
           '32 inclusive.')
    try:
        ivalue=int(value)
    except:
        raise argparse.ArgumentTypeError(error)
    if ivalue<2 or ivalue>32:
        raise argparse.ArgumentTypeError(error)
    return ivalue

def parse_arguments():
    """Function to parse command line arguments.

    Uses argparse.ArgumentParser objet to parse script arguments.
    For argument list see script documentation.

    Returns:
        A Namespace object with the arguments.
    """

    parser=argparse.ArgumentParser() #Create an ArgumentParser
    parser.add_argument('image',
                        help='The image filename that will be used as input.')
    parser.add_argument('-w','--width',type=int,
                        help=('Force scaling to the width provided. It '
                              'is ignored if width is greater than original '
                              'width.'))
    parser.add_argument('-c','--colors',type=number_of_colors,default=32,
                        help=('Recolor the image with the specified number of '
                              'colors (including white). 32 standard colors '
                              'are used if omitted. Valid values are between '
                              '2 and 32 inclusive).'))
    parser.add_argument('-l','--lang',default='en',
                        help=('Language to be used for the legend in the output'
                              ' spreadsheet. Default is \'en\' (English).'))
    parser.add_argument('-r','--resample',type=str.upper,default='BICUBIC',
                        choices=FILTERS,
                        help=('Resample filter to be used while resizing. It '
                              'is ignored if resizing width is not set or '
                              'if an invalid option given. Available options: '
                              'NEAREST, BOX, BILINEAR, HAMMING, BICUBIC or '
                              'LANCZOS. Use NEAREST if image is already in'
                              'pixel art form. Default is BICUBIC.'))
    return parser.parse_args() #Do the parsing
    
def main():
    """Main script function.

    This function parses command line, processes image, loads language file and
    creates spreadsheets by calling other script functions.

    Returns:
        0 for succesfull completion.
        1 if error occured while reading image file.
        2 if error occured while writing workbook file.
    """
    
    args=parse_arguments() #Parse the arguments
    #Load and process image
    img=process_image(args.image,args.width,args.colors,args.resample)
    if not img: #Exit if error occured while processing image
        return 1
    captions,colornames=load_language(args.lang) #Load language file
    wb=create_workbook(args.image+'.xlsx',img,captions,colornames)
    if not wb: #Exit with error code if saving failed
        return 2
    return 0 #Success!!!
        
if __name__ == "__main__":
    sys.exit(main()) #Run the script and exit
